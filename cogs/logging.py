# logging.py - ПОЛНЫЙ ИСПРАВЛЕННЫЙ КОД
import discord
from discord.ext import commands
from discord import ui
from utils.config import CHANNELS, COLORS, ROLES
from utils.database import Database
import datetime
import io
import aiosqlite

# Для Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️ openpyxl не установлен. Установите: pip install openpyxl")

db = Database()

class AddMemberLogModal(ui.Modal, title='➕ Добавить информацию'):
    def __init__(self):
        super().__init__(timeout=300)
    
    user_mention = ui.TextInput(
        label='Пользователь (тег или ID)',
        placeholder='@user или 123456789',
        required=True,
        max_length=100
    )
    
    nickname = ui.TextInput(
        label='Игровой никнейм',
        placeholder='Например: John_Smith',
        required=True,
        max_length=100
    )
    
    passport_phone = ui.TextInput(
        label='Паспорт / Телефон',
        placeholder='Например: 245313 / 4113048',
        required=False,
        max_length=100
    )
    
    additional_info = ui.TextInput(
        label='Дополнительная информация',
        placeholder='Игровое время, часовой пояс, реальное имя и т.д.',
        style=discord.TextStyle.paragraph,
        required=False,
        max_length=500
    )
    
    notes = ui.TextInput(
        label='Заметки',
        placeholder='Любая важная информация...',
        style=discord.TextStyle.paragraph,
        required=False,
        max_length=1000
    )

    async def on_submit(self, interaction: discord.Interaction):
        try:
            # Парсим User ID
            user_input = self.user_mention.value.strip()
            user_id = None
            
            if user_input.startswith('<@') and user_input.endswith('>'):
                user_id = int(user_input.strip('<@!>'))
            else:
                try:
                    user_id = int(user_input)
                except:
                    await interaction.response.send_message("❌ Неверный формат User ID!", ephemeral=True)
                    return
            
            # Парсим паспорт и телефон
            passport = "Не указано"
            phone = "Не указано"
            
            if self.passport_phone.value:
                parts = [p.strip() for p in self.passport_phone.value.split('/')]
                if len(parts) >= 1:
                    passport = parts[0]
                if len(parts) >= 2:
                    phone = parts[1]
            
            # Сохраняем
            success = await save_member_log(
                user_id,
                self.nickname.value,
                passport,
                phone,
                self.additional_info.value or "Нет",
                self.notes.value or "Нет",
                str(interaction.user)
            )
            
            if success:
                embed = discord.Embed(
                    title="✅ Информация добавлена",
                    description=f"Данные об участнике **{self.nickname.value}** успешно сохранены!",
                    color=COLORS["SUCCESS"]
                )
                embed.add_field(name="User ID", value=f"`{user_id}`", inline=True)
                embed.add_field(name="Никнейм", value=f"`{self.nickname.value}`", inline=True)
                await interaction.response.send_message(embed=embed, ephemeral=True)
            else:
                await interaction.response.send_message("❌ Не удалось сохранить информацию", ephemeral=True)
                
        except Exception as e:
            print(f"❌ Ошибка при добавлении информации: {e}")
            await interaction.response.send_message(f"❌ Ошибка: {str(e)}", ephemeral=True)

async def init_member_logs_db():
    """Инициализация таблицы для логов участников"""
    try:
        db_path = db.db_path
        async with aiosqlite.connect(db_path) as database:
            await database.execute('''
                CREATE TABLE IF NOT EXISTS member_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    nickname TEXT NOT NULL,
                    passport TEXT,
                    phone TEXT,
                    additional_info TEXT,
                    notes TEXT,
                    added_by TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            await database.commit()
            print("✅ Таблица member_logs создана/проверена")
            return True
    except Exception as e:
        print(f"❌ Ошибка инициализации таблицы member_logs: {e}")
        return False

async def save_member_log(user_id: int, nickname: str, passport: str, phone: str, 
                          additional_info: str, notes: str, added_by: str):
    """Сохраняет информацию об участнике"""
    try:
        db_path = db.db_path
        async with aiosqlite.connect(db_path) as database:
            await database.execute('''
                INSERT INTO member_logs (user_id, nickname, passport, phone, additional_info, notes, added_by)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (user_id, nickname, passport, phone, additional_info, notes, added_by))
            await database.commit()
            print(f"✅ Информация об участнике {nickname} сохранена")
            return True
    except Exception as e:
        print(f"❌ Ошибка сохранения информации: {e}")
        return False

async def get_all_member_logs():
    """Получает все логи участников"""
    try:
        db_path = db.db_path
        async with aiosqlite.connect(db_path) as database:
            cursor = await database.execute('SELECT * FROM member_logs ORDER BY created_at DESC')
            results = await cursor.fetchall()
            return results
    except Exception as e:
        print(f"❌ Ошибка получения логов: {e}")
        return []

async def create_excel_file():
    """Создает Excel файл со всеми логами участников"""
    if not EXCEL_AVAILABLE:
        return None
    
    try:
        # Получаем все логи
        logs = await get_all_member_logs()
        
        if not logs:
            return None
        
        # Создаем новую рабочую книгу
        wb = Workbook()
        ws = wb.active
        ws.title = "Логи участников"
        
        # Стили
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовки
        headers = ['№', 'User ID', 'Никнейм', 'Паспорт', 'Телефон', 'Доп. информация', 'Заметки', 'Добавил', 'Дата добавления']
        ws.append(headers)
        
        # Стилизация заголовков
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Заполняем данные
        for idx, log in enumerate(logs, start=1):
            # log: (id, user_id, nickname, passport, phone, additional_info, notes, added_by, created_at)
            row = [
                idx,
                log[1],  # user_id
                log[2],  # nickname
                log[3],  # passport
                log[4],  # phone
                log[5],  # additional_info
                log[6],  # notes
                log[7],  # added_by
                log[8]   # created_at
            ]
            ws.append(row)
            
            # Стилизация строк
            for cell in ws[idx + 1]:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Автоширина колонок
        column_widths = [5, 20, 25, 15, 15, 30, 40, 20, 20]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # Сохраняем в BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
        
    except Exception as e:
        print(f"❌ Ошибка создания Excel файла: {e}")
        return None

class AddLogButtonView(ui.View):
    """View с кнопкой для открытия формы добавления лога"""
    def __init__(self):
        super().__init__(timeout=300)
    
    @ui.button(label='📝 Открыть форму', style=discord.ButtonStyle.primary)
    async def open_form(self, interaction: discord.Interaction, button: ui.Button):
        await interaction.response.send_modal(AddMemberLogModal())

class Logging(commands.Cog):
    def __init__(self, bot):
        self.bot = bot

    @commands.Cog.listener()
    async def on_ready(self):
        # Инициализируем таблицу для логов участников
        await init_member_logs_db()

    async def log_action(self, action: str, description: str, color: int, guild: discord.Guild, **kwargs):
        """Универсальная функция для логирования"""
        log_channel = guild.get_channel(CHANNELS["LOG_CHANNEL"])
        if not log_channel:
            return

        embed = discord.Embed(
            title=f"📝 {action}",
            description=description,
            color=color,
            timestamp=datetime.datetime.utcnow()
        )

        for key, value in kwargs.items():
            if value:
                embed.add_field(name=key.replace('_', ' ').title(), value=value, inline=True)

        embed.set_footer(text=f"Server: {guild.name}")
        await log_channel.send(embed=embed)

    # ========== КОМАНДЫ ДЛЯ РАБОТЫ С ЛОГАМИ ==========

    @commands.command(name="add_log", description="Добавить информацию об участнике")
    @commands.has_any_role(ROLES["OWNER"])
    async def add_log(self, ctx):
        """Добавить информацию об участнике (только для Owner)"""
        try:
            embed = discord.Embed(
                title="➕ Добавить лог участника",
                description="Нажмите кнопку ниже, чтобы открыть форму добавления информации",
                color=COLORS["INFO"]
            )
            
            view = AddLogButtonView()
            await ctx.send(embed=embed, view=view)
            
        except Exception as e:
            print(f"❌ Ошибка в add_log: {e}")
            await ctx.send(f"❌ Ошибка: {str(e)}")

    @commands.command(name="get_logs", description="Скачать все логи участников в Excel")
    @commands.has_any_role(ROLES["OWNER"])
    async def get_logs(self, ctx):
        """Скачать Excel файл со всеми логами (только для Owner)"""
        try:
            if not EXCEL_AVAILABLE:
                await ctx.send("❌ Библиотека openpyxl не установлена! Установите: `pip install openpyxl`")
                return
            
            msg = await ctx.send("⏳ Создаю Excel файл...")
            
            # Создаем Excel файл
            excel_file = await create_excel_file()
            
            if not excel_file:
                await msg.edit(content="❌ Нет данных для экспорта или произошла ошибка!")
                return
            
            # Создаем Discord файл
            current_date = datetime.datetime.now().strftime("%d.%m.%Y_%H-%M")
            discord_file = discord.File(
                excel_file,
                filename=f"Ocean_Logs_{current_date}.xlsx"
            )
            
            embed = discord.Embed(
                title="📊 Логи участников",
                description="Excel файл со всеми логами участников готов!",
                color=COLORS["SUCCESS"],
                timestamp=datetime.datetime.utcnow()
            )
            
            logs = await get_all_member_logs()
            embed.add_field(name="Всего записей", value=f"`{len(logs)}`", inline=True)
            embed.add_field(name="Дата экспорта", value=f"`{current_date}`", inline=True)
            
            await msg.delete()
            await ctx.send(embed=embed, file=discord_file)
            
        except Exception as e:
            print(f"❌ Ошибка при создании Excel файла: {e}")
            await ctx.send(f"❌ Ошибка при создании файла: {str(e)}")

    @commands.command(name="show_logs", description="Посмотреть последние логи участников")
    @commands.has_any_role(ROLES["OWNER"])
    async def show_logs(self, ctx, limit: int = 10):
        """Показать последние записи логов (только для Owner)"""
        try:
            logs = await get_all_member_logs()
            
            if not logs:
                await ctx.send("📋 Логи участников пусты")
                return
            
            # Ограничиваем количество
            display_logs = logs[:min(limit, len(logs))]
            
            embed = discord.Embed(
                title="📋 Логи участников",
                description=f"Показано последних записей: {len(display_logs)} из {len(logs)}",
                color=COLORS["INFO"]
            )
            
            for log in display_logs:
                # log: (id, user_id, nickname, passport, phone, additional_info, notes, added_by, created_at)
                field_value = (
                    f"**User ID:** `{log[1]}`\n"
                    f"**Паспорт:** `{log[3]}`\n"
                    f"**Телефон:** `{log[4]}`\n"
                    f"**Доп. инфо:** {log[5]}\n"
                    f"**Заметки:** {log[6]}\n"
                    f"**Добавил:** {log[7]}"
                )
                embed.add_field(
                    name=f"👤 {log[2]}",
                    value=field_value,
                    inline=False
                )
            
            await ctx.send(embed=embed)
            
        except Exception as e:
            print(f"❌ Ошибка при просмотре логов: {e}")
            await ctx.send(f"❌ Ошибка: {str(e)}")
    
    @commands.command(name="test_logs")
    @commands.has_any_role(ROLES["OWNER"])
    async def test_logs(self, ctx):
        """Тестовая команда для проверки работы"""
        await ctx.send("✅ Команда работает! Система логов активна.", delete_after=10)

    # ========== СТАНДАРТНОЕ ЛОГИРОВАНИЕ ==========

    @commands.Cog.listener()
    async def on_message_delete(self, message):
        if message.author.bot:
            return
        
        await self.log_action(
            "Сообщение удалено",
            f"Сообщение от {message.author.mention} было удалено в {message.channel.mention}",
            COLORS["ERROR"],
            message.guild,
            author=message.author.mention,
            channel=message.channel.mention,
            content=message.content or "Вложение/Embed"
        )

    @commands.Cog.listener()
    async def on_message_edit(self, before, after):
        if before.author.bot or before.content == after.content:
            return
        
        await self.log_action(
            "Сообщение изменено",
            f"Сообщение от {before.author.mention} было изменено в {before.channel.mention}",
            COLORS["WARNING"],
            before.guild,
            author=before.author.mention,
            channel=before.channel.mention,
            before=before.content[:1024] if before.content else "Вложение/Embed",
            after=after.content[:1024] if after.content else "Вложение/Embed"
        )

    @commands.Cog.listener()
    async def on_member_join(self, member):
        await self.log_action(
            "Участник присоединился",
            f"{member.mention} присоединился к серверу",
            COLORS["SUCCESS"],
            member.guild,
            user=member.mention,
            account_created=discord.utils.format_dt(member.created_at, 'R')
        )

    @commands.Cog.listener()
    async def on_member_remove(self, member):
        await self.log_action(
            "Участник покинул сервер",
            f"{member.mention} покинул сервер",
            COLORS["ERROR"],
            member.guild,
            user=member.mention,
            joined_at=discord.utils.format_dt(member.joined_at, 'R') if member.joined_at else "Неизвестно"
        )

    @commands.Cog.listener()
    async def on_member_update(self, before, after):
        if before.roles != after.roles:
            added_roles = [role for role in after.roles if role not in before.roles]
            removed_roles = [role for role in before.roles if role not in after.roles]

            if added_roles:
                await self.log_action(
                    "Роль добавлена",
                    f"Пользователю {after.mention} добавлены роли",
                    COLORS["SUCCESS"],
                    after.guild,
                    user=after.mention,
                    roles=", ".join([role.mention for role in added_roles])
                )

            if removed_roles:
                await self.log_action(
                    "Роль удалена",
                    f"У пользователя {after.mention} удалены роли",
                    COLORS["ERROR"],
                    after.guild,
                    user=after.mention,
                    roles=", ".join([role.mention for role in removed_roles])
                )

    @commands.Cog.listener()
    async def on_voice_state_update(self, member, before, after):
        if before.channel != after.channel:
            if not before.channel and after.channel:
                await self.log_action(
                    "Подключение к голосовому каналу",
                    f"{member.mention} подключился к голосовому каналу",
                    COLORS["SUCCESS"],
                    member.guild,
                    user=member.mention,
                    channel=after.channel.mention
                )
            elif before.channel and not after.channel:
                await self.log_action(
                    "Отключение от голосового канала",
                    f"{member.mention} отключился от голосового канала",
                    COLORS["ERROR"],
                    member.guild,
                    user=member.mention,
                    channel=before.channel.mention
                )
            elif before.channel and after.channel:
                await self.log_action(
                    "Перемещение между голосовыми каналами",
                    f"{member.mention} переместился между голосовыми каналами",
                    COLORS["WARNING"],
                    member.guild,
                    user=member.mention,
                    from_channel=before.channel.mention,
                    to_channel=after.channel.mention
                )

    @commands.Cog.listener()
    async def on_guild_role_create(self, role):
        await self.log_action(
            "Роль создана",
            f"Создана новая роль {role.mention}",
            COLORS["SUCCESS"],
            role.guild,
            role=role.mention,
            color=str(role.color)
        )

    @commands.Cog.listener()
    async def on_guild_role_delete(self, role):
        await self.log_action(
            "Роль удалена",
            f"Роль **{role.name}** была удалена",
            COLORS["ERROR"],
            role.guild,
            role=role.name
        )

    @commands.Cog.listener()
    async def on_guild_role_update(self, before, after):
        if before.name != after.name:
            await self.log_action(
                "Роль переименована",
                f"Роль была переименована",
                COLORS["WARNING"],
                after.guild,
                before_name=before.name,
                after_name=after.name
            )

    @commands.Cog.listener()
    async def on_guild_channel_create(self, channel):
        await self.log_action(
            "Канал создан",
            f"Создан новый канал {channel.mention}",
            COLORS["SUCCESS"],
            channel.guild,
            channel=channel.mention,
            type=channel.type.name
        )

    @commands.Cog.listener()
    async def on_guild_channel_delete(self, channel):
        await self.log_action(
            "Канал удален",
            f"Канал **{channel.name}** был удален",
            COLORS["ERROR"],
            channel.guild,
            channel=channel.name,
            type=channel.type.name
        )

async def setup(bot):
    await bot.add_cog(Logging(bot))