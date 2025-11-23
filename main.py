# main.py - ИСПРАВЛЕННЫЙ
import discord
from discord.ext import commands, tasks
import os
import asyncio
from dotenv import load_dotenv
from utils.database import Database
import random

# Загружаем переменные из .env файла
load_dotenv()

# Настройки бота - ИЗМЕНЕН ПРЕФИКС на !
intents = discord.Intents.all()
bot = commands.Bot(command_prefix="!", intents=intents, help_command=None)
db = Database()

# Информация о создателе
BOT_CREATOR = "Ocean Family"
BOT_VERSION = "4.0"

@bot.event
async def on_ready():
    print(f'🌊 Ocean Bot {bot.user} успешно запущен!')
    print(f'📊 Подключен к {len(bot.guilds)} серверам:')
    
    for guild in bot.guilds:
        print(f'   - {guild.name} (ID: {guild.id})')
    
    # Инициализация базы данных
    try:
        await db.init_db()
        print("✅ База данных инициализирована")
    except Exception as e:
        print(f"❌ Ошибка инициализации базы данных: {e}")
    
    # Небольшая задержка для стабилизации
    await asyncio.sleep(2)
    
    # Загрузка модулей
    loaded_cogs = 0
    for filename in os.listdir("./cogs"):
        if filename.endswith(".py"):
            try:
                await bot.load_extension(f"cogs.{filename[:-3]}")
                print(f'✅ Загружен модуль: {filename[:-3]}')
                loaded_cogs += 1
            except Exception as e:
                print(f'❌ Ошибка загрузки {filename}: {e}')
    
    print(f'📦 Всего загружено модулей: {loaded_cogs}')
    
    try:
        synced = await bot.tree.sync()
        print(f'✅ Синхронизировано {len(synced)} slash команд:')
        for cmd in synced:
            print(f'   - /{cmd.name}')
    except Exception as e:
        print(f'❌ Ошибка синхронизации команд: {e}')
    
    # Печатаем обычные команды
    print(f'✅ Доступно {len(bot.commands)} обычных команд:')
    for cmd in bot.commands:
        print(f'   - !{cmd.name}')
    
    # Дополнительная задержка для регистрации view
    await asyncio.sleep(1)
    print("✅ Все view зарегистрированы")
    
    # Запускаем динамический статус
    if not change_status.is_running():
        change_status.start()
        print("🎮 Динамический статус запущен")
    
    print("🎉 Бот полностью готов к работе!")
    print("💡 Используйте префикс ! для обычных команд (например: !add_log)")
    print("💡 Используйте / для slash команд (например: /ping)")

@tasks.loop(seconds=15)  # Меняется каждые 15 секунд
async def change_status():
    """Динамическая смена статуса бота"""
    try:
        statuses = [
            discord.Activity(type=discord.ActivityType.playing, name="Ocean Family 🌊"),
            discord.Activity(type=discord.ActivityType.watching, name="за участниками"),
            discord.Activity(type=discord.ActivityType.listening, name="команды"),
            discord.Activity(type=discord.ActivityType.playing, name="с контрактами"),
            discord.Activity(type=discord.ActivityType.watching, name="дни рождения"),
        ]
        activity = random.choice(statuses)
        await bot.change_presence(activity=activity, status=discord.Status.online)
    except Exception as e:
        print(f"❌ Ошибка при смене статуса: {e}")

@change_status.before_loop
async def before_change_status():
    """Ждем, пока бот будет готов, перед запуском смены статуса"""
    await bot.wait_until_ready()

@bot.event
async def on_guild_join(guild):
    """При подключении к новому серверу"""
    print(f'✅ Бот добавлен на сервер: {guild.name} (ID: {guild.id})')
    
    # Ищем канал для отправки приветственного сообщения
    system_channel = guild.system_channel
    if system_channel and system_channel.permissions_for(guild.me).send_messages:
        embed = discord.Embed(
            title="🌊 Ocean Bot подключен!",
            description="Спасибо за добавление бота Ocean Family!",
            color=0x00ffff
        )
        embed.add_field(
            name="📝 Основные команды:",
            value="• `!setup_apply` - Настроить форму для вступления\n• `!setup_contracts` - Настроить контракты\n• `!add_log` - Добавить лог участника",
            inline=False
        )
        embed.add_field(
            name="⚙️ Настройка:",
            value="Не забудьте настроить ID ролей и каналов в config.py",
            inline=False
        )
        embed.set_footer(text="Ocean Family")
        
        await system_channel.send(embed=embed)

@bot.event
async def on_command_error(ctx, error):
    """Обработка ошибок команд"""
    if isinstance(error, commands.CommandNotFound):
        return
    elif isinstance(error, commands.MissingPermissions):
        await ctx.send("❌ У вас недостаточно прав для выполнения этой команды.")
    elif isinstance(error, commands.MissingAnyRole):
        await ctx.send("❌ У вас нет необходимой роли для выполнения этой команды.")
    elif isinstance(error, commands.MissingRequiredArgument):
        await ctx.send(f"❌ Не хватает аргумента: {error.param.name}")
    elif isinstance(error, commands.BadArgument):
        await ctx.send("❌ Неверный аргумент команды.")
    else:
        print(f"❌ Необработанная ошибка команды: {error}")
        import traceback
        traceback.print_exception(type(error), error, error.__traceback__)

@bot.hybrid_command(name="ping", description="Проверить задержку бота")
async def ping(ctx):
    """Проверка пинга бота"""
    latency = round(bot.latency * 1000)
    
    embed = discord.Embed(
        title="🏓 Понг!",
        description=f"Задержка бота: **{latency}мс**",
        color=0x00ff00
    )
    
    await ctx.send(embed=embed)

@bot.hybrid_command(name="bot_info", description="Информация о боте")
async def bot_info(ctx):
    """Информация о боте"""
    embed = discord.Embed(
        title="🌊 Ocean Bot Information",
        color=0x00ffff
    )
    
    embed.add_field(name="👑 Создатель", value=BOT_CREATOR, inline=True)
    embed.add_field(name="📚 Версия", value=BOT_VERSION, inline=True)
    embed.add_field(name="🏓 Задержка", value=f"{round(bot.latency * 1000)}мс", inline=True)
    
    embed.add_field(
        name="📊 Статистика",
        value=f"**Серверов:** {len(bot.guilds)}\n**Пользователей:** {len(bot.users)}",
        inline=False
    )
    
    embed.add_field(
        name="🛠️ Функции",
        value="• Система заявок\n• Контракты\n• Дни рождения\n• Логирование участников\n• Приветствия",
        inline=False
    )
    
    embed.add_field(
        name="🎮 Статус",
        value="Динамический статус включен ✅",
        inline=False
    )
    
    embed.set_footer(text=f"Ocean Family Bot | Created by {BOT_CREATOR}")
    
    await ctx.send(embed=embed)

@bot.hybrid_command(name="reload", description="Перезагрузить модули бота (только для владельцев)")
@commands.is_owner()
async def reload(ctx):
    """Перезагрузка модулей бота"""
    reloaded = 0
    for filename in os.listdir("./cogs"):
        if filename.endswith(".py"):
            try:
                await bot.reload_extension(f"cogs.{filename[:-3]}")
                reloaded += 1
            except Exception as e:
                await ctx.send(f"❌ Ошибка перезагрузки {filename}: {e}")
                return
    
    embed = discord.Embed(
        title="✅ Перезагрузка завершена",
        description=f"Успешно перезагружено {reloaded} модулей",
        color=0x00ff00
    )
    await ctx.send(embed=embed)

@bot.hybrid_command(name="clear_data", description="Очистить все данные (только для владельцев)")
@commands.is_owner()
async def clear_data(ctx):
    """Очистка всех данных (для тестирования)"""
    try:
        await db.clear_applications()
        await db.clear_birthdays()
        await db.clear_contracts()
        
        embed = discord.Embed(
            title="✅ Данные очищены",
            description="Все данные базы данных были очищены",
            color=0x00ff00
        )
        await ctx.send(embed=embed)
    except Exception as e:
        embed = discord.Embed(
            title="❌ Ошибка очистки",
            description=f"Ошибка: {e}",
            color=0xff0000
        )
        await ctx.send(embed=embed)

@bot.hybrid_command(name="status_toggle", description="Включить/выключить динамический статус")
@commands.is_owner()
async def status_toggle(ctx):
    """Управление динамическим статусом"""
    if change_status.is_running():
        change_status.stop()
        embed = discord.Embed(
            title="⏸️ Статус отключен",
            description="Динамический статус выключен",
            color=0xffff00
        )
    else:
        change_status.start()
        embed = discord.Embed(
            title="▶️ Статус включен",
            description="Динамический статус включен",
            color=0x00ff00
        )
    
    await ctx.send(embed=embed)

async def main():
    """Основная функция запуска бота"""
    try:
        # Загружаем переменные окружения из .env файла
        load_dotenv()
        
        # Получаем токен из переменной окружения
        token = os.getenv('DISCORD_BOT_TOKEN')
        
        if not token:
            print("❌ Ошибка: Токен не найден в переменных окружения.")
            print("💡 Создайте файл .env с содержанием: DISCORD_BOT_TOKEN=ваш_токен")
            return
        
        print("🚀 Запуск Ocean Bot...")
        await bot.start(token)
        
    except KeyboardInterrupt:
        print("\n🛑 Остановка бота...")
        await bot.close()
    except Exception as e:
        print(f"❌ Критическая ошибка: {e}")
        await bot.close()

if __name__ == "__main__":
    asyncio.run(main())# logging.py - РАСШИРЕННАЯ СИСТЕМА ЛОГИРОВАНИЯ С EXCEL
import discord
from discord.ext import commands
from discord import ui
from utils.config import CHANNELS, COLORS, ROLES
from utils.database import Database
import datetime
import io
import aiosqlite

# Для Excel понадобится установить: pip install openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️ openpyxl не установлен. Установите: pip install openpyxl")

db = Database()

class AddMemberLogModal(ui.Modal, title='➕ Добавить информацию'):
    def __init__(self):
        super().__init__(timeout=300)
    
    user_mention = ui.TextInput(
        label='Пользователь (тег или ID)',
        placeholder='@user или 123456789',
        required=True,
        max_length=100
    )
    
    nickname = ui.TextInput(
        label='Игровой никнейм',
        placeholder='Например: John_Smith',
        required=True,
        max_length=100
    )
    
    passport_phone = ui.TextInput(
        label='Паспорт / Телефон',
        placeholder='Например: 245313 / 4113048',
        required=False,
        max_length=100
    )
    
    additional_info = ui.TextInput(
        label='Дополнительная информация',
        placeholder='Игровое время, часовой пояс, реальное имя и т.д.',
        style=discord.TextStyle.paragraph,
        required=False,
        max_length=500
    )
    
    notes = ui.TextInput(
        label='Заметки',
        placeholder='Любая важная информация...',
        style=discord.TextStyle.paragraph,
        required=False,
        max_length=1000
    )

    async def on_submit(self, interaction: discord.Interaction):
        try:
            # Парсим User ID из тега или ID
            user_input = self.user_mention.value.strip()
            user_id = None
            
            # Пробуем извлечь ID из упоминания <@123456789>
            if user_input.startswith('<@') and user_input.endswith('>'):
                user_id = int(user_input.strip('<@!>'))
            else:
                # Пробуем напрямую как число
                try:
                    user_id = int(user_input)
                except:
                    await interaction.response.send_message("❌ Неверный формат User ID!", ephemeral=True)
                    return
            
            # Парсим паспорт и телефон
            passport = "Не указано"
            phone = "Не указано"
            
            if self.passport_phone.value:
                parts = [p.strip() for p in self.passport_phone.value.split('/')]
                if len(parts) >= 1:
                    passport = parts[0]
                if len(parts) >= 2:
                    phone = parts[1]
            
            # Сохраняем в базу данных
            success = await save_member_log(
                user_id,
                self.nickname.value,
                passport,
                phone,
                self.additional_info.value or "Нет",
                self.notes.value or "Нет",
                str(interaction.user)
            )
            
            if success:
                embed = discord.Embed(
                    title="✅ Информация добавлена",
                    description=f"Данные об участнике **{self.nickname.value}** успешно сохранены!",
                    color=COLORS["SUCCESS"]
                )
                embed.add_field(name="User ID", value=f"`{user_id}`", inline=True)
                embed.add_field(name="Никнейм", value=f"`{self.nickname.value}`", inline=True)
                await interaction.response.send_message(embed=embed, ephemeral=True)
            else:
                await interaction.response.send_message("❌ Не удалось сохранить информацию", ephemeral=True)
                
        except Exception as e:
            print(f"❌ Ошибка при добавлении информации: {e}")
            await interaction.response.send_message(f"❌ Ошибка: {str(e)}", ephemeral=True)

async def init_member_logs_db():
    """Инициализация таблицы для логов участников"""
    try:
        db_path = db.db_path
        async with aiosqlite.connect(db_path) as database:
            await database.execute('''
                CREATE TABLE IF NOT EXISTS member_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    nickname TEXT NOT NULL,
                    passport TEXT,
                    phone TEXT,
                    additional_info TEXT,
                    notes TEXT,
                    added_by TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            await database.commit()
            print("✅ Таблица member_logs создана/проверена")
            return True
    except Exception as e:
        print(f"❌ Ошибка инициализации таблицы member_logs: {e}")
        return False

async def save_member_log(user_id: int, nickname: str, passport: str, phone: str, 
                          additional_info: str, notes: str, added_by: str):
    """Сохраняет информацию об участнике"""
    try:
        db_path = db.db_path
        async with aiosqlite.connect(db_path) as database:
            await database.execute('''
                INSERT INTO member_logs (user_id, nickname, passport, phone, additional_info, notes, added_by)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (user_id, nickname, passport, phone, additional_info, notes, added_by))
            await database.commit()
            print(f"✅ Информация об участнике {nickname} сохранена")
            return True
    except Exception as e:
        print(f"❌ Ошибка сохранения информации: {e}")
        return False

async def get_all_member_logs():
    """Получает все логи участников"""
    try:
        db_path = db.db_path
        async with aiosqlite.connect(db_path) as database:
            cursor = await database.execute('SELECT * FROM member_logs ORDER BY created_at DESC')
            results = await cursor.fetchall()
            return results
    except Exception as e:
        print(f"❌ Ошибка получения логов: {e}")
        return []

async def create_excel_file():
    """Создает Excel файл со всеми логами участников"""
    if not EXCEL_AVAILABLE:
        return None
    
    try:
        # Получаем все логи
        logs = await get_all_member_logs()
        
        if not logs:
            return None
        
        # Создаем новую рабочую книгу
        wb = Workbook()
        ws = wb.active
        ws.title = "Логи участников"
        
        # Стили
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовки
        headers = ['№', 'User ID', 'Никнейм', 'Паспорт', 'Телефон', 'Доп. информация', 'Заметки', 'Добавил', 'Дата добавления']
        ws.append(headers)
        
        # Стилизация заголовков
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Заполняем данные
        for idx, log in enumerate(logs, start=1):
            # log: (id, user_id, nickname, passport, phone, additional_info, notes, added_by, created_at)
            row = [
                idx,
                log[1],  # user_id
                log[2],  # nickname
                log[3],  # passport
                log[4],  # phone
                log[5],  # additional_info
                log[6],  # notes
                log[7],  # added_by
                log[8]   # created_at
            ]
            ws.append(row)
            
            # Стилизация строк
            for cell in ws[idx + 1]:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Автоширина колонок
        column_widths = [5, 20, 25, 15, 15, 30, 40, 20, 20]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # Сохраняем в BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
        
    except Exception as e:
        print(f"❌ Ошибка создания Excel файла: {e}")
        return None

class Logging(commands.Cog):
    def __init__(self, bot):
        self.bot = bot

    @commands.Cog.listener()
    async def on_ready(self):
        # Инициализируем таблицу для логов участников
        await init_member_logs_db()

    async def log_action(self, action: str, description: str, color: int, guild: discord.Guild, **kwargs):
        """Универсальная функция для логирования"""
        log_channel = guild.get_channel(CHANNELS["LOG_CHANNEL"])
        if not log_channel:
            return

        embed = discord.Embed(
            title=f"📝 {action}",
            description=description,
            color=color,
            timestamp=datetime.datetime.utcnow()
        )

        for key, value in kwargs.items():
            if value:
                embed.add_field(name=key.replace('_', ' ').title(), value=value, inline=True)

        embed.set_footer(text=f"Server: {guild.name}")
        await log_channel.send(embed=embed)

    # ========== КОМАНДЫ ДЛЯ РАБОТЫ С ЛОГАМИ ==========

    @commands.command(name="add_log", description="Добавить информацию об участнике")
    @commands.has_any_role(ROLES["OWNER"])
    async def add_log(self, ctx):
        """Добавить информацию об участнике (только для Owner)"""
        try:
            # Создаем embed с инструкцией и кнопкой
            embed = discord.Embed(
                title="➕ Добавить лог участника",
                description="Нажмите кнопку ниже, чтобы открыть форму добавления информации",
                color=COLORS["INFO"]
            )
            
            # Создаем view с кнопкой
            view = AddLogButtonView()
            await ctx.send(embed=embed, view=view)
            
        except Exception as e:
            print(f"❌ Ошибка в add_log: {e}")
            await ctx.send(f"❌ Ошибка: {str(e)}")

    @commands.command(name="get_logs", description="Скачать все логи участников в Excel")
    @commands.has_any_role(ROLES["OWNER"])
    async def get_logs(self, ctx):
        """Скачать Excel файл со всеми логами (только для Owner)"""
        try:
            if not EXCEL_AVAILABLE:
                await ctx.send("❌ Библиотека openpyxl не установлена! Установите: `pip install openpyxl`")
                return
            
            msg = await ctx.send("⏳ Создаю Excel файл...")
            
            # Создаем Excel файл
            excel_file = await create_excel_file()
            
            if not excel_file:
                await msg.edit(content="❌ Нет данных для экспорта или произошла ошибка!")
                return
            
            # Создаем Discord файл
            current_date = datetime.datetime.now().strftime("%d.%m.%Y_%H-%M")
            discord_file = discord.File(
                excel_file,
                filename=f"Ocean_Logs_{current_date}.xlsx"
            )
            
            embed = discord.Embed(
                title="📊 Логи участников",
                description="Excel файл со всеми логами участников готов!",
                color=COLORS["SUCCESS"],
                timestamp=datetime.datetime.utcnow()
            )
            
            logs = await get_all_member_logs()
            embed.add_field(name="Всего записей", value=f"`{len(logs)}`", inline=True)
            embed.add_field(name="Дата экспорта", value=f"`{current_date}`", inline=True)
            
            await msg.delete()
            await ctx.send(embed=embed, file=discord_file)
            
        except Exception as e:
            print(f"❌ Ошибка при создании Excel файла: {e}")
            await ctx.send(f"❌ Ошибка при создании файла: {str(e)}")

    @commands.command(name="show_logs", description="Посмотреть последние логи участников")
    @commands.has_any_role(ROLES["OWNER"])
    async def show_logs(self, ctx, limit: int = 10):
        """Показать последние записи логов (только для Owner)"""
        try:
            logs = await get_all_member_logs()
            
            if not logs:
                await ctx.send("📋 Логи участников пусты")
                return
            
            # Ограничиваем количество
            display_logs = logs[:min(limit, len(logs))]
            
            embed = discord.Embed(
                title="📋 Логи участников",
                description=f"Показано последних записей: {len(display_logs)} из {len(logs)}",
                color=COLORS["INFO"]
            )
            
            for log in display_logs:
                # log: (id, user_id, nickname, passport, phone, additional_info, notes, added_by, created_at)
                field_value = (
                    f"**User ID:** `{log[1]}`\n"
                    f"**Паспорт:** `{log[3]}`\n"
                    f"**Телефон:** `{log[4]}`\n"
                    f"**Доп. инфо:** {log[5]}\n"
                    f"**Заметки:** {log[6]}\n"
                    f"**Добавил:** {log[7]}"
                )
                embed.add_field(
                    name=f"👤 {log[2]}",
                    value=field_value,
                    inline=False
                )
            
            await ctx.send(embed=embed)
            
        except Exception as e:
            print(f"❌ Ошибка при просмотре логов: {e}")
            await ctx.send(f"❌ Ошибка: {str(e)}")
    
    @commands.command(name="test_logs")
    @commands.has_any_role(ROLES["OWNER"])
    async def test_logs(self, ctx):
        """Тестовая команда для проверки работы"""
        await ctx.send("✅ Команда работает! Система логов активна.", delete_after=10)

class AddLogButtonView(ui.View):
    """View с кнопкой для открытия формы добавления лога"""
    def __init__(self):
        super().__init__(timeout=300)
    
    @ui.button(label='📝 Открыть форму', style=discord.ButtonStyle.primary)
    async def open_form(self, interaction: discord.Interaction, button: ui.Button):
        await interaction.response.send_modal(AddMemberLogModal())

    # ========== СТАНДАРТНОЕ ЛОГИРОВАНИЕ ==========

    @commands.Cog.listener()
    async def on_message_delete(self, message):
        if message.author.bot:
            return
        
        await self.log_action(
            "Сообщение удалено",
            f"Сообщение от {message.author.mention} было удалено в {message.channel.mention}",
            COLORS["ERROR"],
            message.guild,
            author=message.author.mention,
            channel=message.channel.mention,
            content=message.content or "Вложение/Embed"
        )

    @commands.Cog.listener()
    async def on_message_edit(self, before, after):
        if before.author.bot or before.content == after.content:
            return
        
        await self.log_action(
            "Сообщение изменено",
            f"Сообщение от {before.author.mention} было изменено в {before.channel.mention}",
            COLORS["WARNING"],
            before.guild,
            author=before.author.mention,
            channel=before.channel.mention,
            before=before.content[:1024] if before.content else "Вложение/Embed",
            after=after.content[:1024] if after.content else "Вложение/Embed"
        )

    @commands.Cog.listener()
    async def on_member_join(self, member):
        await self.log_action(
            "Участник присоединился",
            f"{member.mention} присоединился к серверу",
            COLORS["SUCCESS"],
            member.guild,
            user=member.mention,
            account_created=discord.utils.format_dt(member.created_at, 'R')
        )

    @commands.Cog.listener()
    async def on_member_remove(self, member):
        await self.log_action(
            "Участник покинул сервер",
            f"{member.mention} покинул сервер",
            COLORS["ERROR"],
            member.guild,
            user=member.mention,
            joined_at=discord.utils.format_dt(member.joined_at, 'R') if member.joined_at else "Неизвестно"
        )

    @commands.Cog.listener()
    async def on_member_update(self, before, after):
        if before.roles != after.roles:
            added_roles = [role for role in after.roles if role not in before.roles]
            removed_roles = [role for role in before.roles if role not in after.roles]

            if added_roles:
                await self.log_action(
                    "Роль добавлена",
                    f"Пользователю {after.mention} добавлены роли",
                    COLORS["SUCCESS"],
                    after.guild,
                    user=after.mention,
                    roles=", ".join([role.mention for role in added_roles])
                )

            if removed_roles:
                await self.log_action(
                    "Роль удалена",
                    f"У пользователя {after.mention} удалены роли",
                    COLORS["ERROR"],
                    after.guild,
                    user=after.mention,
                    roles=", ".join([role.mention for role in removed_roles])
                )

    @commands.Cog.listener()
    async def on_voice_state_update(self, member, before, after):
        if before.channel != after.channel:
            if not before.channel and after.channel:
                await self.log_action(
                    "Подключение к голосовому каналу",
                    f"{member.mention} подключился к голосовому каналу",
                    COLORS["SUCCESS"],
                    member.guild,
                    user=member.mention,
                    channel=after.channel.mention
                )
            elif before.channel and not after.channel:
                await self.log_action(
                    "Отключение от голосового канала",
                    f"{member.mention} отключился от голосового канала",
                    COLORS["ERROR"],
                    member.guild,
                    user=member.mention,
                    channel=before.channel.mention
                )
            elif before.channel and after.channel:
                await self.log_action(
                    "Перемещение между голосовыми каналами",
                    f"{member.mention} переместился между голосовыми каналами",
                    COLORS["WARNING"],
                    member.guild,
                    user=member.mention,
                    from_channel=before.channel.mention,
                    to_channel=after.channel.mention
                )

    @commands.Cog.listener()
    async def on_guild_role_create(self, role):
        await self.log_action(
            "Роль создана",
            f"Создана новая роль {role.mention}",
            COLORS["SUCCESS"],
            role.guild,
            role=role.mention,
            color=str(role.color)
        )

    @commands.Cog.listener()
    async def on_guild_role_delete(self, role):
        await self.log_action(
            "Роль удалена",
            f"Роль **{role.name}** была удалена",
            COLORS["ERROR"],
            role.guild,
            role=role.name
        )

    @commands.Cog.listener()
    async def on_guild_role_update(self, before, after):
        if before.name != after.name:
            await self.log_action(
                "Роль переименована",
                f"Роль была переименована",
                COLORS["WARNING"],
                after.guild,
                before_name=before.name,
                after_name=after.name
            )

    @commands.Cog.listener()
    async def on_guild_channel_create(self, channel):
        await self.log_action(
            "Канал создан",
            f"Создан новый канал {channel.mention}",
            COLORS["SUCCESS"],
            channel.guild,
            channel=channel.mention,
            type=channel.type.name
        )

    @commands.Cog.listener()
    async def on_guild_channel_delete(self, channel):
        await self.log_action(
            "Канал удален",
            f"Канал **{channel.name}** был удален",
            COLORS["ERROR"],
            channel.guild,
            channel=channel.name,
            type=channel.type.name
        )

async def setup(bot):
    await bot.add_cog(Logging(bot))